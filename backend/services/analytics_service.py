# backend/services/analytics_service.py
"""
Analytics Service - API endpoints for data analysis and visualization
"""
from fastapi import APIRouter, HTTPException
from pydantic import BaseModel
from typing import Optional
from backend.analytics.data_logger import DataLogger
import os
import csv
from pathlib import Path

router = APIRouter(prefix="/api/analytics", tags=["Analytics"])

logger = DataLogger(log_dir="logs")


# Request Models
class LogEntry(BaseModel):
    emotion: str
    smile: str
    eyes: str
    posture: str
    cognitive_state: Optional[str] = ""
    mood: Optional[str] = ""
    sentiment: Optional[float] = 0.0
    environment_feedback: Optional[str] = ""


# Endpoints
@router.get("/summary")
def get_analytics_summary():
    """
    Get overall analytics summary from CSV logs.
    """
    try:
        log_file = Path("tests/logs/syntwin_log.csv")
        
        if not log_file.exists():
            return {
                "success": False,
                "message": "No log file found. Run detection first."
            }
        
        with open(log_file, 'r') as f:
            reader = csv.DictReader(f)
            data = list(reader)
        
        if not data:
            return {
                "success": False,
                "message": "Log file is empty"
            }
        
        # Calculate statistics
        emotions = [row['emotion'] for row in data if row.get('emotion')]
        postures = [row['posture'] for row in data if row.get('posture')]
        sentiments = [float(row['sentiment']) for row in data if row.get('sentiment') and row['sentiment']]
        
        from collections import Counter
        
        emotion_counts = Counter(emotions)
        posture_counts = Counter(postures)
        
        avg_sentiment = sum(sentiments) / len(sentiments) if sentiments else 0
        
        return {
            "success": True,
            "data": {
                "total_entries": len(data),
                "emotion_distribution": dict(emotion_counts.most_common(5)),
                "posture_distribution": dict(posture_counts.most_common(5)),
                "average_sentiment": round(avg_sentiment, 2),
                "sentiment_trend": "Positive" if avg_sentiment > 0.3 else "Negative" if avg_sentiment < -0.3 else "Neutral"
            }
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/timeline")
def get_analytics_timeline(limit: int = 50):
    """
    Get timeline of analytics data.
    
    Query params:
    - limit: Number of entries to return (default: 50)
    """
    try:
        log_file = Path("tests/logs/syntwin_log.csv")
        
        if not log_file.exists():
            return {
                "success": False,
                "message": "No log file found"
            }
        
        with open(log_file, 'r') as f:
            reader = csv.DictReader(f)
            data = list(reader)
        
        # Get last N entries
        timeline = data[-limit:] if len(data) > limit else data
        
        return {
            "success": True,
            "count": len(timeline),
            "data": timeline
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/log")
def log_analytics_data(entry: LogEntry):
    """
    Log an analytics entry to CSV.
    
    Body: LogEntry with emotion, smile, eyes, posture, etc.
    """
    try:
        entry_dict = entry.dict()
        logger.log_entry(entry_dict)
        
        return {
            "success": True,
            "message": "Data logged successfully",
            "data": entry_dict
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/emotion-trends")
def get_emotion_trends(hours: int = 24):
    """
    Get emotion trends over time.
    
    Query params:
    - hours: Time period in hours (default: 24)
    """
    try:
        log_file = Path("tests/logs/syntwin_log.csv")
        
        if not log_file.exists():
            return {
                "success": False,
                "message": "No log file found"
            }
        
        with open(log_file, 'r') as f:
            reader = csv.DictReader(f)
            data = list(reader)
        
        # Group by hour
        from datetime import datetime, timedelta
        from collections import defaultdict
        
        time_threshold = datetime.now() - timedelta(hours=hours)
        hourly_emotions = defaultdict(list)
        
        for row in data:
            if row.get('timestamp') and row.get('emotion'):
                try:
                    timestamp = datetime.strptime(row['timestamp'], "%Y-%m-%d %H:%M:%S")
                    if timestamp >= time_threshold:
                        hour_key = timestamp.strftime("%Y-%m-%d %H:00")
                        hourly_emotions[hour_key].append(row['emotion'])
                except:
                    continue
        
        # Calculate dominant emotion per hour
        trends = []
        for hour, emotions in sorted(hourly_emotions.items()):
            from collections import Counter
            emotion_counter = Counter(emotions)
            dominant = emotion_counter.most_common(1)[0] if emotion_counter else ("Unknown", 0)
            
            trends.append({
                "hour": hour,
                "dominant_emotion": dominant[0],
                "count": dominant[1],
                "total_detections": len(emotions)
            })
        
        return {
            "success": True,
            "data": trends
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/clear-logs")
def clear_analytics_logs():
    """
    Clear CSV log file.
    """
    try:
        logger.clear_logs()
        return {
            "success": True,
            "message": "Analytics logs cleared"
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/export-excel")
def export_excel_report():
    """
    Generate and download Excel report with all analytics data.
    """
    try:
        from fastapi.responses import StreamingResponse
        from io import BytesIO
        from datetime import datetime
        import sqlite3
        
        # Try to import openpyxl, if not available use csv
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            use_excel = True
        except ImportError:
            use_excel = False
        
        # Get data from database
        db_path = Path(__file__).parent.parent / "database" / "syntwin.db"
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        
        # Fetch all detection data
        cursor.execute("""
            SELECT timestamp, emotion, posture, sentiment
            FROM detector_logs
            ORDER BY timestamp DESC
        """)
        detections = cursor.fetchall()
        
        # Fetch statistics
        cursor.execute("""
            SELECT 
                COUNT(*) as total,
                AVG(sentiment) as avg_sentiment,
                emotion,
                COUNT(*) as emotion_count
            FROM detector_logs
            GROUP BY emotion
        """)
        emotion_stats = cursor.fetchall()
        
        cursor.execute("""
            SELECT 
                posture,
                COUNT(*) as posture_count
            FROM detector_logs
            GROUP BY posture
        """)
        posture_stats = cursor.fetchall()
        
        conn.close()
        
        if use_excel:
            # Create Excel workbook
            wb = Workbook()
            
            # Summary Sheet
            ws_summary = wb.active
            ws_summary.title = "Summary"
            
            # Header styling
            header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            ws_summary['A1'] = 'SynTwin Detection Report'
            ws_summary['A1'].font = Font(bold=True, size=16)
            ws_summary['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
            
            ws_summary['A4'] = 'Total Detections'
            ws_summary['B4'] = len(detections)
            
            # Emotion Distribution
            ws_summary['A6'] = 'Emotion Distribution'
            ws_summary['A6'].font = header_font
            ws_summary['A6'].fill = header_fill
            ws_summary['A7'] = 'Emotion'
            ws_summary['B7'] = 'Count'
            ws_summary['A7'].font = header_font
            ws_summary['B7'].font = header_font
            
            row = 8
            for stat in emotion_stats:
                ws_summary[f'A{row}'] = stat[2]  # emotion
                ws_summary[f'B{row}'] = stat[3]  # count
                row += 1
            
            # Posture Distribution
            row += 1
            ws_summary[f'A{row}'] = 'Posture Distribution'
            ws_summary[f'A{row}'].font = header_font
            ws_summary[f'A{row}'].fill = header_fill
            row += 1
            ws_summary[f'A{row}'] = 'Posture'
            ws_summary[f'B{row}'] = 'Count'
            ws_summary[f'A{row}'].font = header_font
            ws_summary[f'B{row}'].font = header_font
            row += 1
            
            for stat in posture_stats:
                ws_summary[f'A{row}'] = stat[0]  # posture
                ws_summary[f'B{row}'] = stat[1]  # count
                row += 1
            
            # Detection Details Sheet
            ws_details = wb.create_sheet("Detection Details")
            ws_details['A1'] = 'Timestamp'
            ws_details['B1'] = 'Emotion'
            ws_details['C1'] = 'Posture'
            ws_details['D1'] = 'Sentiment'
            
            for col in ['A1', 'B1', 'C1', 'D1']:
                ws_details[col].font = header_font
                ws_details[col].fill = header_fill
            
            for idx, det in enumerate(detections, start=2):
                ws_details[f'A{idx}'] = det[0]
                ws_details[f'B{idx}'] = det[1]
                ws_details[f'C{idx}'] = det[2]
                ws_details[f'D{idx}'] = det[3]
            
            # Auto-adjust column widths
            for ws in [ws_summary, ws_details]:
                for column in ws.columns:
                    max_length = 0
                    column = list(column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column[0].column_letter].width = adjusted_width
            
            # Save to BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            filename = f"syntwin_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f'attachment; filename="{filename}"',
                    "Access-Control-Expose-Headers": "Content-Disposition"
                }
            )
        else:
            # Fallback to CSV if openpyxl not available
            import io
            text_output = io.StringIO()
            writer = csv.writer(text_output)
            
            writer.writerow(['Timestamp', 'Emotion', 'Posture', 'Sentiment'])
            for det in detections:
                writer.writerow(det)
            
            output = BytesIO(text_output.getvalue().encode('utf-8'))
            output.seek(0)
            
            filename = f"syntwin_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            
            return StreamingResponse(
                output,
                media_type="text/csv",
                headers={
                    "Content-Disposition": f'attachment; filename="{filename}"',
                    "Access-Control-Expose-Headers": "Content-Disposition"
                }
            )
            
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/health")
def health_check():
    """Check if analytics service is running."""
    return {
        "success": True,
        "service": "Analytics",
        "status": "running"
    }
